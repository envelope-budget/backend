import csv
from datetime import datetime

from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_transactions_csv(transactions, start_date, end_date):
    """Export transactions to CSV format."""
    response = HttpResponse(content_type="text/csv")
    filename = f"spending_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Write header
    writer.writerow(["Date", "Payee", "Account", "Envelope", "Amount", "Memo"])

    # Write transactions
    for transaction in transactions:
        writer.writerow(
            [
                transaction.date.strftime("%Y-%m-%d"),
                transaction.payee.name if transaction.payee else "",
                transaction.account.name,
                transaction.envelope.name if transaction.envelope else "",
                f"{transaction.amount / 1000:.2f}",
                transaction.memo or "",
            ]
        )

    return response


def export_transactions_xlsx(transactions, start_date, end_date, budget_name):
    """Export transactions to Excel format."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        )

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spending Report"

    # Add title and metadata
    ws["A1"] = f"Spending Report - {budget_name}"
    ws["A2"] = (
        f"Period: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
    )
    ws["A3"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"

    # Style the title
    title_font = Font(size=16, bold=True)
    ws["A1"].font = title_font

    # Add headers starting at row 5
    headers = ["Date", "Payee", "Account", "Envelope", "Amount", "Memo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Add transaction data
    for row, transaction in enumerate(transactions, 6):
        ws.cell(row=row, column=1, value=transaction.date.strftime("%Y-%m-%d"))
        ws.cell(
            row=row, column=2, value=transaction.payee.name if transaction.payee else ""
        )
        ws.cell(row=row, column=3, value=transaction.account.name)
        ws.cell(
            row=row,
            column=4,
            value=transaction.envelope.name if transaction.envelope else "",
        )
        ws.cell(row=row, column=5, value=f"{transaction.amount / 1000:.2f}")
        ws.cell(row=row, column=6, value=transaction.memo or "")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, ValueError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"spending_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def export_transactions_markdown(transactions, start_date, end_date, budget_name):
    """Export transactions to Markdown format."""
    response = HttpResponse(content_type="text/markdown")
    filename = f"spending_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.md"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Calculate totals
    total_income = sum(t.amount for t in transactions if t.amount > 0) / 1000
    total_spent = abs(sum(t.amount for t in transactions if t.amount < 0)) / 1000
    net_amount = (sum(t.amount for t in transactions)) / 1000

    # Generate markdown content
    content = f"""# Spending Report - {budget_name}

**Period:** {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}
**Generated:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

## Summary

| Metric | Amount |
|--------|--------|
| Total Income | ${total_income:,.2f} |
| Total Spent | ${total_spent:,.2f} |
| Net Amount | ${net_amount:,.2f} |

## Transactions ({len(transactions)} total)

| Date | Payee | Account | Envelope | Amount | Memo |
|------|-------|---------|----------|--------|------|
"""

    for transaction in transactions:
        payee = transaction.payee.name if transaction.payee else "—"
        envelope = transaction.envelope.name if transaction.envelope else "—"
        memo = transaction.memo or "—"
        amount = f"${transaction.amount / 1000:.2f}"

        # Escape pipe characters in content
        payee = payee.replace("|", "\\|")
        envelope = envelope.replace("|", "\\|")
        memo = memo.replace("|", "\\|")

        content += f"| {transaction.date.strftime('%Y-%m-%d')} | {payee} | {transaction.account.name} | {envelope} | {amount} | {memo} |\n"

    response.write(content)
    return response


def export_budget_csv(budget_data, budget_name):
    """Export budget data as CSV"""
    response = HttpResponse(content_type="text/csv")
    filename = f"budget_report_{budget_name}_{datetime.now().strftime('%Y%m%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(["Category", "Envelope", "Monthly Budget Amount", "Notes"])

    for item in budget_data:
        writer.writerow(
            [
                item["category_name"],
                item["envelope_name"],
                f"${item['monthly_budget_amount']:.2f}",
                item["note"],
            ]
        )

    return response


def export_budget_xlsx(budget_data, budget_name):
    """Export budget data as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        )

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Report"

    # Add headers
    headers = ["Category", "Envelope", "Monthly Budget Amount", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Add data
    for row, item in enumerate(budget_data, 2):
        ws.cell(row=row, column=1, value=item["category_name"])
        ws.cell(row=row, column=2, value=item["envelope_name"])
        ws.cell(row=row, column=3, value=item["monthly_budget_amount"])
        ws.cell(row=row, column=4, value=item["note"])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"budget_report_{budget_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def export_budget_markdown(budget_data, budget_name):
    """Export budget data as Markdown"""
    response = HttpResponse(content_type="text/markdown")
    filename = f"budget_report_{budget_name}_{datetime.now().strftime('%Y%m%d')}.md"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    content = f"# Budget Report - {budget_name}\n\n"
    content += f"Generated on: {datetime.now().strftime('%B %d, %Y')}\n\n"
    content += "| Category | Envelope | Monthly Budget Amount | Notes |\n"
    content += "|----------|----------|----------------------|-------|\n"

    for item in budget_data:
        content += f"| {item['category_name']} | {item['envelope_name']} | ${item['monthly_budget_amount']:.2f} | {item['note']} |\n"

    response.write(content)
    return response


def export_spending_by_category_csv(spending_data, budget_name, start_date, end_date):
    """Export spending by category data as CSV"""
    response = HttpResponse(content_type="text/csv")
    filename = f"spending_by_category_{budget_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Get months list from first item (if exists)
    months_list = spending_data[0].get("months_list", []) if spending_data else []

    # Create header row with monthly columns FIRST
    header = ["Category", "Envelope"]

    # Add monthly columns first
    for month in months_list:
        header.append(month["name"])

    # Then add the summary columns
    header.extend(["Total Spent", "Average Spent", "Budget Amount", "Notes"])
    writer.writerow(header)

    for item in spending_data:
        row = [
            item["category_name"],
            item["envelope_name"],
        ]

        # Add monthly spending data first
        monthly_spending = item.get("monthly_spending", {})
        for month in months_list:
            amount = monthly_spending.get(month["name"], 0)
            row.append(f"${amount:.2f}")

        # Then add summary data
        row.extend(
            [
                f"${item['total_spent']:.2f}",
                f"${item['average_spent']:.2f}",
                f"${item['budget_amount']:.2f}",
                item["note"],
            ]
        )
        writer.writerow(row)

    return response


def export_spending_by_category_xlsx(spending_data, budget_name, start_date, end_date):
    """Export spending by category data as Excel file"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        )

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Spending by Category"

    # Add title and metadata
    ws["A1"] = f"Spending by Category Report - {budget_name}"
    ws["A2"] = f"Period: {start_date.strftime('%B %Y')} - {end_date.strftime('%B %Y')}"
    ws["A3"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"

    # Style the title
    title_font = Font(size=16, bold=True)
    ws["A1"].font = title_font

    # Get months list from first item (if exists)
    months_list = spending_data[0].get("months_list", []) if spending_data else []

    # Add headers starting at row 5 with monthly columns FIRST
    headers = ["Category", "Envelope"]

    # Add monthly columns first
    for month in months_list:
        headers.append(month["name"])

    # Then add summary columns
    headers.extend(["Total Spent", "Average Spent", "Budget Amount", "Notes"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Add data
    for row, item in enumerate(spending_data, 6):
        col = 1
        ws.cell(row=row, column=col, value=item["category_name"])
        col += 1
        ws.cell(row=row, column=col, value=item["envelope_name"])
        col += 1

        # Add monthly spending data first
        monthly_spending = item.get("monthly_spending", {})
        for month in months_list:
            amount = monthly_spending.get(month["name"], 0)
            ws.cell(row=row, column=col, value=amount)
            col += 1

        # Then add summary data
        ws.cell(row=row, column=col, value=item["total_spent"])
        col += 1
        ws.cell(row=row, column=col, value=item["average_spent"])
        col += 1
        ws.cell(row=row, column=col, value=item["budget_amount"])
        col += 1
        ws.cell(row=row, column=col, value=item["note"])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"spending_by_category_{budget_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


def export_spending_by_category_markdown(
    spending_data, budget_name, start_date, end_date
):
    """Export spending by category data as Markdown"""
    response = HttpResponse(content_type="text/markdown")
    filename = f"spending_by_category_{budget_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.md"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Calculate totals
    total_spent = sum(item["total_spent"] for item in spending_data)
    total_budget = sum(item["budget_amount"] for item in spending_data)

    # Get months list from first item (if exists)
    months_list = spending_data[0].get("months_list", []) if spending_data else []

    content = f"""# Spending by Category Report - {budget_name}

**Period:** {start_date.strftime('%B %Y')} - {end_date.strftime('%B %Y')}
**Generated:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}

## Summary

| Metric | Amount |
|--------|--------|
| Total Spent | ${total_spent:,.2f} |
| Total Budget | ${total_budget:,.2f} |
| Budget Variance | ${total_budget - total_spent:,.2f} |

## Spending by Category ({len(spending_data)} envelopes)

| Category | Envelope |"""

    # Add monthly column headers first
    for month in months_list:
        content += f" {month['name']} |"

    # Then add summary column headers
    content += " Total Spent | Average Spent | Budget Amount | Notes |\n"
    content += "|----------|----------|"

    # Add monthly column separators first
    for month in months_list:
        content += "---------------|"

    # Then add summary column separators
    content += "-------------|---------------|---------------|-------|\n"

    for item in spending_data:
        # Escape pipe characters in content
        category = item["category_name"].replace("|", "\\|")
        envelope = item["envelope_name"].replace("|", "\\|")
        note = item["note"].replace("|", "\\|") if item["note"] else "—"

        content += f"| {category} | {envelope} |"

        # Add monthly spending data first
        monthly_spending = item.get("monthly_spending", {})
        for month in months_list:
            amount = monthly_spending.get(month["name"], 0)
            content += f" ${amount:.2f} |"

        # Then add summary data
        content += f" ${item['total_spent']:.2f} | ${item['average_spent']:.2f} | ${item['budget_amount']:.2f} | {note} |\n"

    response.write(content)
    return response
